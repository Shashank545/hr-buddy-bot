"""
This script is used to extract knowledge items from xlsx to csv.
"""
import csv
from pathlib import Path
from openpyxl import load_workbook


PATH_TO_EXCEL_FILE = "data/Reference Knowledges.xlsx"
WORKSHEET_NAMES = [
    "Red Book",
    "Insights",
    "IFRS FAQ",
]
FIELDS_TO_EXTRACT = [
    "ナレッジID",
    "データ種別大分類",
    "詳細・回答内容",
]

FIELD_NAMES_MAP_FROM_EN_TO_JA = {
    "ナレッジID": "knowledge_id",
    "データ種別大分類": "big_data_type",
    "詳細・回答内容": "answer",
}

OUTPUT_DIR = "data/csv"

def main():
    """Extract knowledge items to CSV file."""
    path_to_xlsx_file = Path(PATH_TO_EXCEL_FILE)
    if not path_to_xlsx_file.is_file():
        print(f"[ERROR] Make sure to create file '{PATH_TO_EXCEL_FILE}'")
        exit(1)

    workbook = load_workbook(path_to_xlsx_file)
    for worksheet in workbook.worksheets:
        if worksheet.title in WORKSHEET_NAMES:
            header_idx = dict()
            rows = []
            for row in worksheet.iter_rows():
                if not header_idx:
                    # Process the first row
                    for idx, cell in enumerate(row):
                        if cell.value in FIELDS_TO_EXTRACT:
                            header_idx[cell.value] = idx
                    continue

                cells = [cell.value for cell in row]
                rows.append({
                    FIELD_NAMES_MAP_FROM_EN_TO_JA[field]: cells[idx]
                    for field, idx in header_idx.items()
                })

            if rows:
                output_filename = Path(f"{OUTPUT_DIR}/{worksheet.title.replace(' ', '_').lower()}.csv")
                with open(output_filename, mode="w", encoding="utf-8", newline="") as file_:
                    dwriter = csv.DictWriter(
                        file_,
                        fieldnames=FIELD_NAMES_MAP_FROM_EN_TO_JA.values(),
                        delimiter=",",
                        quotechar='"',
                        quoting=csv.QUOTE_ALL
                    )
                    dwriter.writeheader()
                    dwriter.writerows(rows)
                print(f"[INFO] '{len(rows)}' documents are saved at '{output_filename}'")


if __name__ == "__main__":
    main()
