"""Util methods."""
import json
import re
from csv import DictWriter, QUOTE_ALL
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook


class SearchOption(str, Enum):
    """Search options"""
    BM25 = "bm25"
    Semantic = "semantic"
    Vector = "vector"
    VectorBM25 = "vector-bm25"
    VectorSemantic = "vector-semantic"

SEARCH_OPTIONS = list(map(lambda c: c.value, SearchOption))
SEARCH_OPTIONS_VECTOR = {SearchOption.Vector, SearchOption.VectorBM25, SearchOption.VectorSemantic}
MAP_FROM_SEARCH_OPTION_TO_INT = dict(map(lambda c: (c[1].value, c[0]), enumerate(SearchOption)))

class ApproachOption(str, Enum):
    """Approach options"""
    Approach1= "rr"
    Approach2 = "rrrr"
    Approach3 = "rrr"
    Approach4 = "rrrt"

MAP_FROM_INT_TO_APPROACH = dict(map(lambda c: (c[0]+1, c[1].value), enumerate(ApproachOption)))
APPROACH_OPTIONS = set(MAP_FROM_INT_TO_APPROACH.keys())

class LLMOption(str, Enum):
    """LLM options"""
    Gpt35Turbo = "gpt-35-turbo"
    Gpt4 = "gpt-4"

LLM_OPTIONS = list(map(lambda c: c.value, LLMOption))

# The following values may need to be updated if the input file format changes
PATH_TO_EXCEL_FILE = "data/ChatKOMEI PoC評価報告書.xlsx"
WORKSHEET_NAME = "Result (template)"
MIN_ROW = 5  # This is the first that is a test case
COL_INDEX_QUESTION = 1
COL_INDEX_ANSWER = 3
COL_INDEX_KNOWLEDGE_ID = 5
KNOWLEDGE_ID_REGEX = re.compile(r"^Acifrs\-\d{6}$")


def clean_text(text: str) -> str:
    """Clean up text."""
    text = text.replace("\n", "")
    return text


def extract_test_cases(max_count: int = None) -> list:
    """Extract test cases from 'PATH_TO_EXCEL_FILE'."""
    path_to_xlsx_file = Path(PATH_TO_EXCEL_FILE)
    if not path_to_xlsx_file.is_file():
        print(f"[ERROR] Make sure to create file '{PATH_TO_EXCEL_FILE}'")
        exit(1)

    workbook = load_workbook(path_to_xlsx_file)
    rows = []
    for row_idx, row in enumerate(workbook[WORKSHEET_NAME].iter_rows(min_row=MIN_ROW)):
        if isinstance(max_count, int) and row_idx >= max_count:
            break

        cells = [cell.value for cell in row]
        rows.append({
            "id": row_idx + 1,
            "query": clean_text(cells[COL_INDEX_QUESTION]),
            "answer": clean_text(cells[COL_INDEX_ANSWER]),
            "knowledge_id": {
                item
                for item in cells[COL_INDEX_KNOWLEDGE_ID].split("\n")
                if re.match(KNOWLEDGE_ID_REGEX, item)
            },
        })
    
    if rows:
        return rows
    else:
        print("[INFO] No test cases to run.")
        exit(1)


def to_csv(data: list, output_filename: Path):
    """Save data to CSV file."""
    with open(output_filename, mode="w", encoding="utf_8_sig") as file_:
        dict_writer = DictWriter(file_, fieldnames=data[0].keys(), quoting=QUOTE_ALL)
        dict_writer.writeheader()
        dict_writer.writerows(data)


def to_json(data: list, output_filename: Path):
    """Save items into JSON file."""
    with open(output_filename, mode="w", encoding="utf-8") as file_:
        json.dump(data, file_, indent=2, ensure_ascii=False)
